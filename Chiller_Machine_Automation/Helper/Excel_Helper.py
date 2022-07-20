#*************************************************************************************************
# Standard and Project libraries
#*************************************************************************************************
import sys
sys.path.append("../")

import openpyxl
from openpyxl import load_workbook
import sys
import os
import time

from Configuration import config


#*************************************************************************************************
#Class Name     : ExcelOperation()
#Description    : This class is used to perform the excel operations
#Parameters     : NA
#*************************************************************************************************
class ExcelOperation:
    
    def __init__(self,excelFileName,DefaultSheetName=None):
        self.excelFile = excelFileName
        self.sheet = DefaultSheetName
        
        self.activeSheet = None
        self.openWorkBook = None
        self.openSheet = None
        
        self.workBook = None
        
    #*************************************************************************************************
    #Function Name  : checkExcel()
    #Description    : This function is used to check whether the required excel file exist or not
    #Parameters     : NA
    #*************************************************************************************************
    def checkExcel(self):
        path = config.OUTPUT_PATH
        os.chdir(path)
        currentPath = os.getcwd()
        if self.excelFile in os.listdir(currentPath):
            self.openExcel()
        else:
            self.createExcel()
        
    #*************************************************************************************************
    #Function Name  : createExcel()
    #Description    : This function is used to create a new excel file
    #Parameters     : NA
    #*************************************************************************************************
    def createExcel(self):
        try:
            self.workBook = openpyxl.Workbook()
            self.activeSheet = self.workBook.active
            self.writeHeadings()
        except:
            print ("Warning: There is some problem with Excel file. Please check that the excel file is available in Input folder")
            sys.exit()
            
    #*************************************************************************************************
    #Function Name  : writeHeadings()
    #Description    : This function is used to write the heading in newly created excel file
    #Parameters     : NA
    #*************************************************************************************************
    def writeHeadings(self):
        Heading1 = ["Sr. No.","Input Parameters","Chilled Water Parameters","Cooling Water Parameters","Outputs","Measured Parameters"]
        
        for ele in Heading1:
            if ele == "Sr. No.":
                self.activeSheet.merge_cells("A1:A2")
                self.activeSheet.cell(row=1,column=1).value=ele
            elif ele == "Input Parameters":
                self.activeSheet.merge_cells("B1:F1")
                self.activeSheet.cell(row=1,column=2).value=ele
            elif ele == "Chilled Water Parameters":
                self.activeSheet.merge_cells("G1:I1")
                self.activeSheet.cell(row=1,column=7).value=ele
            elif ele == "Cooling Water Parameters":
                self.activeSheet.merge_cells("J1:L1")
                self.activeSheet.cell(row=1,column=10).value=ele
            elif ele == "Outputs":
                self.activeSheet.merge_cells("M1:P1")
                self.activeSheet.cell(row=1,column=13).value=ele
            elif ele == "Measured Parameters":
                self.activeSheet.merge_cells("Q1:T1")
                self.activeSheet.cell(row=1,column=17).value=ele

        self.saveExcel()
        
    #*************************************************************************************************
    #Function Name  : openExcel()
    #Description    : This function is used to open an existing excel file
    #Parameters     : NA
    #*************************************************************************************************
    def openExcel(self):
        path = config.INPUT_PATH
        os.chdir(path)
        
        try:
            self.openWorkBook = load_workbook(self.excelFile)
        except:
            print ("Warning in openExcel(): There is some problem with Excel file. Please check that the excel file is available in Input folder")
            sys.exit()
    
    #*************************************************************************************************
    #Function Name  : createSheet()
    #Description    : This function is used to create a sheet
    #Parameters     : [ sheet Name ]
    #*************************************************************************************************
    def createSheet(self,SheetName):
        try:
            sheet = self.excelFile.create_sheet(index=0, title=SheetName)
        except:
            print ("Warning: There is some problem with excel sheet "+str(SheetName))
            sys.exit()
            
    
    #*************************************************************************************************
    #Function Name  : getAllSheets()
    #Description    : This function is used to list down all the sheets name
    #Parameters     : NA
    #*************************************************************************************************        
    def getAllSheets(self):
        try:
            AllSheets = self.excelFile.get_sheet_names()
        except:
            print ("Warning: There is some problem with excel")
            sys.exit()
            
        return AllSheets
    
    
    #*************************************************************************************************
    #Function Name  : getSheet()
    #Description    : This function is used to access one particular sheet
    #Parameters     : [ sheet Name ]
    #*************************************************************************************************       
    def getSheet(self,SheetName=None):
        try:
            #self.openSheet = self.openWorkBook.get_sheet_by_name(SheetName)
            self.openSheet = self.openWorkBook.active
        except:
            print ("Warning: There is some problem with excel sheet "+SheetName)
            sys.exit()
            
    #*************************************************************************************************
    #Function Name  : writeWorkSheet()
    #Description    : This function is used to write the data in excel
    #Parameters     : [ row, column, value ]
    #*************************************************************************************************  
    def writeWorkSheet(self,row_,clm_,value_):
            self.openSheet.cell(row=row_,column=clm_).value=value_

    #*************************************************************************************************
    #Function Name  : activeWorksheet()
    #Description    : This function is used to active the worksheet
    #Parameters     : []
    #*************************************************************************************************
    def activeWorksheet(self):
        try:
            ActiveSheet = self.excelFile.active
        except:
            print ("Warning: There is some problem with excel")
            sys.exit()
            
        return ActiveSheet
    
    
    #*************************************************************************************************
    #Function Name  : saveExcel()
    #Description    : This function is used to save the Excel file
    #Parameters     : []
    #*************************************************************************************************
    def saveExcel(self):
        path = config.OUTPUT_PATH
        os.chdir(path)
        currentPath = os.getcwd()
        self.openWorkBook.save(currentPath+config.OUTPUT_EXCEL)

    #*************************************************************************************************
    #Function Name  : RemoveExistingFileIfAvailable()
    #Description    : This function is used to remove an existing Excel file
    #Parameters     : []
    #*************************************************************************************************    
    def RemoveExistingFileIfAvailable(self):
        path = config.OUTPUT_PATH
        os.chdir(path)
        currentPath = os.getcwd()
        for file in os.listdir(currentPath):
            if file.endswith('.xlsx') or file.endswith('.xls'):
                os.remove(file)
    
    #*************************************************************************************************
    #Function Name  : fetchEffiency()
    #Description    : This function is used to fetch the Efficiency data and filter the best efficiency
    #Parameters     : []
    #*************************************************************************************************             
    def fetchEffiency(self):
        maxRows = self.openSheet.max_row
        maxColumn = self.openSheet.max_column
        title_ = ["Sr. No.","Chilled Water Flow - (m^3/hr)","Chiller Power - (Kw)","Ambient Dry Build Temp.(DBT) - (Deg C)",
                  "Ambient Wet Build Temp.(WBT) -  (Deg C)", "Relative Humidity (RH) - (%)","Chilled Water Temp. Inlet (Tc in) - (Deg C)",
                  "Chilled Water Temp. Outlet (Tc out) - (Deg C)","Refrigerent Temp. in Chiller (Deg C)",
                  "Cooling Water Temp. Inlet (Tcw in) - (Deg C)","Cooling Water Temp. Outlet (Tc out) - (Deg C)",
                  "Refrigerent Temp. in Chiller (Deg C)","Chiller Range - (Deg C)","Condensor Range - (Deg C)", "Chiller Operating TR - (TR)",
                  "Chiller Power - (Kw)","Chiller Efficiency - (KW/TR)","COP","Chiller Approch - (Deg C)","Condensor Approch - (Deg C)"
                 ]

        EfficiencyData = 0
        RowNumber = 3
        BestEfficiencyConfig = {}
        
        for i in range (3,maxRows+1):
            tempVal = self.openSheet.cell(row=i,column=config.EFFICIENCY_CLM).value
            if tempVal > EfficiencyData:
                EfficiencyData = tempVal
                RowNumber = i
            else:
                pass
        for i in range (1,maxColumn+1):
            tempVal = self.openSheet.cell(row=RowNumber,column=i).value
            para = title_[i-1]
            BestEfficiencyConfig[para]=tempVal
            
        return BestEfficiencyConfig

                
if __name__ == "__main__":
    obj1 = ExcelOperation(config.EXCEL_FILE_TEMPLATE,config.DEFAULT_SHEET)  
    #obj1.checkExcel()     
    obj1.openExcel()
    obj1.getSheet()
    obj1.saveExcel()
    
    